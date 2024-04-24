from pandas import *
from openpyxl import load_workbook
def cgpa(input_excel):
    input_data=load_workbook(input_excel)
    branches_in_input_data=input_data.sheetnames
    objects_list=[]
    for branch in branches_in_input_data:
        if "Analysis" not in branch and branches_in_input_data[-1]=="Overall Analysis":
            branch_wise_data=read_excel(input_data,sheet_name=branch)